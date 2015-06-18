import os
from os import listdir
from os.path import isfile, join

from bs4 import BeautifulSoup
import pandas as pd

from config import *

import openpyxl

class Scraper:
    
    def __init__(self):
        pass

    def scrape(self, path):
    # Takes in the path to an HTML file
    # Returns a list of dicts containing a dataframe for each table in that file
    # And some meta-data about it
    # { file - path scraped, table_number - the nth table in that file, df - pandas DataFrame }
    
        with open(path, "r") as text_file:
            html = text_file.read()
            
        soup = BeautifulSoup(html)
        
        tables = soup.findAll('table')
        
        # This is the list we will be building of dataframes corresponding to the tables in this file
        table_dfs = []
        
        for t, table in enumerate(tables):    
            theads = table.findAll('thead')
            
            if len(theads) == 0:
                table_data = [[cell.text for cell in row(['td','th'])]
                                     for row in table("tr")]
                headers = table_data[0]
                del table_data[0]
            else:
                # In case there is more than one thead, we want the last one
                thead = theads[-1]
                
                # Sometimes there is more than one row in the thead
                # If there is, we want the last one
                thead_rows = thead('tr')
                if len(thead_rows) > 0:
                    thead = thead_rows[-1]
                    
                headers = [cell.text for cell in thead('th')]
                
                tbody = table.findAll('tbody')[0]
                table_data = [[cell.text for cell in row("td")]
                                     for row in tbody("tr")]
            
            
            if len(table_data) > 0:
                
                # Adds dummy headers in case there weren't enough to cover all of the columns in the actual rows
                while len(headers) < max([len(row) for row in table_data]):
                    headers.extend([u'dummy'])
                    
                # Drops headers off the end in case there were too many headers
                headers = headers[0:max([len(row) for row in table_data])]
    
                
                table_dfs.extend([{"file":os.path.basename(path),"table_number":str(t),"df":pd.DataFrame(table_data, columns=headers)}])
        
        # Ensure we free up memory
        soup.decompose()
        
        return table_dfs
    
    def scrape_directory(self, path, verbose=True):
    # Builds a collection of pandas DataFrames along with meta data
    # Corresponding to every table in every HTML file in a directory
        results = []
        

        onlyfiles = [ f for f in listdir(path) if isfile(join(path,f)) ]
        
        for f in onlyfiles:
            if verbose:
                print f
            results.extend(self.scrape(path + f))
            
        return results
    
    def scrape_directory_2013_onwards(self, path, year, verbose=True):
    # Variant of scrape_directory that aligns with the more recent style of data
    # Looks for "year=2012" or equivalent in the filename
        results = []
        
        onlyfiles = [ f for f in listdir(path) if isfile(join(path,f)) ]
        
        for f in onlyfiles:
            if ('year=' + str(year-1)) in f:
                if verbose:
                    print f
                results.extend(self.scrape(path + f))
            
        return results
    
    def do_year(self, year, config, local_save_dir_base):
    # Returns a single dataframe of all data from a single year
        
        # Newer years have a new format
        if year >= 2013:
            results = self.scrape_directory_2013_onwards(local_save_dir_base + config['html_dir'], year, verbose=False)
        else:
            results = self.scrape_directory(local_save_dir_base + config['html_dir'], verbose=False)
            
        # Addenda not used for now
        # results_addenda = filter(lambda item: item['file'] in config['addenda'], results)
        
        results_core = filter(lambda item: item['file'] not in config['core_exclude'], results)
        
        # Rename columns with the broad translators
        for r in results_core:
            for c in r['df'].columns:
                c_stripped = "".join(c.split())
                for rule in stripped_contains_column_translator.keys():
                    if rule in c_stripped:
                        r['df'].rename(columns={c:stripped_contains_column_translator[rule]},inplace=True)
        
        # Rename columns with targeted translators
        for r in results_core:
            for c in r['df'].columns:
                if c not in column_translator.keys():
                    print r
                else:
                    r['df'].rename(columns={c:column_translator[c]},inplace=True)
        
        # Translate filenames into categories. For instance "munic" is "Municipalities"
        for r in results_core:
            for c in category_translator.keys():
                if c in r['file']:
                    r['df']['Category'] = category_translator[c]
                    
        # Useless columns are cleared out
        for r in results_core:
            columns_to_drop = []
            for c in r['df'].columns:
                if c not in core_columns:
                    columns_to_drop.append(c)
            r['df'].drop(columns_to_drop, axis=1, inplace=True)
        
        
        all_dfs = []
        for r in results_core:
            all_dfs.append(r['df'])
        df_core = pd.concat(all_dfs)
        
        # Cleanup Salary
        df_core['Salary'] = df_core['Salary'].str.replace(',','').str.replace('$','')
        df_core['Salary'] = df_core['Salary'].convert_objects(convert_numeric=True)
        
        # Cleanup Benefits
        df_core['Benefits'] = df_core['Benefits'].str.replace(',','').str.replace('$','')
        df_core['Benefits'] = df_core['Benefits'].convert_objects(convert_numeric=True)
        return df_core
    
    def read_2015_excel(self, local_save_dir_base):
        wb = openpyxl.reader.excel.load_workbook(local_save_dir_base + 'pssd_compendium_2014.xlsx', read_only=True)
        
        # Salaries Sheet

        ws = wb.get_sheet_by_name('Salaries')

        category = []
        for row in ws.iter_rows('B2:B111441'):
            for cell in row:
                category.extend([cell.value])

        employer = []
        for row in ws.iter_rows('C2:C111441'):
            for cell in row:
                employer.extend([cell.value])
                
        surname = []
        for row in ws.iter_rows('D2:D111441'):
            for cell in row:
                surname.extend([cell.value])

        given_name = []
        for row in ws.iter_rows('E2:E111441'):
            for cell in row:
                given_name.extend([cell.value])
                
        position = []
        for row in ws.iter_rows('F2:F111441'):
            for cell in row:
                position.extend([cell.value])

        salary = []
        for row in ws.iter_rows('G2:G111441'):
            for cell in row:
                salary.extend([cell.value])
                
        benefits = []
        for row in ws.iter_rows('H2:H111441'):
            for cell in row:
                benefits.extend([cell.value])

        df = pd.DataFrame({'Category':category, 'Employer':employer, 'Surname':surname, 'Given Name':given_name, 'Position':position, 'Salary':salary, 'Benefits':benefits})
        df['year'] = 2015
        df.loc[df['Category'].str.contains('Ministries'),'Category']='Ministries'
        df.loc[df['Category'].str.contains('Colleges'),'Category']='Colleges'
        df.loc[df['Category'].str.contains('Crown'),'Category']='Crown Agencies'
        df.loc[df['Category'].str.contains('Hospitals'),'Category']='Hospitals and Boards of Public Health'
        df.loc[df['Category'].str.contains('Hydro'),'Category']='Hydro One and Ontario Power Generation'
        df.loc[df['Category'].str.contains('Judiciary'),'Category']='Judiciary'
        df.loc[df['Category'].str.contains('Legislative'),'Category']='Legislative Assembly and Offices'
        df.loc[df['Category'].str.contains('Municipalities'),'Category']='Municipalities and Services'
        df.loc[df['Category'].str.contains('Public Sector'),'Category']='Other Public Sector Employers'
        df.loc[df['Category'].str.contains('School'),'Category']='School Boards'
        df.loc[df['Category'].str.contains('Universities'),'Category']='Universities'

        # Seconded Sheet

        ws = wb.get_sheet_by_name('Seconded Employees')

        surname = []
        for row in ws.iter_rows('B2:B146'):
            for cell in row:
                surname.extend([cell.value])

        given_name = []
        for row in ws.iter_rows('C2:C146'):
            for cell in row:
                given_name.extend([cell.value])

        employer = []
        for row in ws.iter_rows('E2:E146'):
            for cell in row:
                employer.extend([cell.value])

        position = []
        for row in ws.iter_rows('F2:F146'):
            for cell in row:
                position.extend([cell.value])

        salary = []
        for row in ws.iter_rows('G2:G146'):
            for cell in row:
                salary.extend([cell.value])
                
        benefits = []
        for row in ws.iter_rows('H2:H146'):
            for cell in row:
                benefits.extend([cell.value])

        df_seconded = pd.DataFrame({'Employer':employer, 'Surname':surname, 'Given Name':given_name, 'Position':position, 'Salary':salary, 'Benefits':benefits})
        df_seconded['year'] = 2015
        df_seconded['Category'] = 'Ministries'

        return pd.concat([df, df_seconded])


    def run(self, local_save_dir_base):
        all_year_dfs=[]
    
        # For years up to 2014 we will scrape the HTML
        for year in configuration.keys():
            print "Processing year " + str(year) + "..."
            df_year = self.do_year(year, configuration[year], local_save_dir_base)
            print "Done."
            
            df_year['year'] = year
            all_year_dfs.append(df_year)
            df_clean = pd.concat(all_year_dfs)

        # For 2015, we process the provided Excel file
        print "Processing year 2015..."
        df_2015 = self.read_2015_excel(local_save_dir_base)
        print "Done."
        df_clean = pd.concat([df_clean, df_2015])

        return df_clean